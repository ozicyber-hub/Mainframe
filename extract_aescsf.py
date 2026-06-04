import openpyxl
import json
import re

XLSX_PATH = r'C:\Users\Alex\Documents\ozireport\V2 AESCSF Toolkit Version V1-1.xlsx'
OUTPUT_PATH = r'C:\Users\Alex\Documents\ozireport\aescsf_v2_extracted.json'

# Domain display names derived from sub-objective titles or manually defined
DOMAIN_NAMES = {
    'ACCESS': 'Identity and Access Management',
    'ARCHITECTURE': 'Cybersecurity Architecture',
    'ASSET': 'Asset, Change, and Configuration Management',
    'PRIVACY': 'Privacy',
    'PROGRAM': 'Cybersecurity Program Management',
    'RESPONSE': 'Incident Response',
    'RISK': 'Risk Management',
    'SITUATION': 'Situational Awareness',
    'THIRD-PARTIES': 'Third-Party Risk Management',
    'THREAT': 'Threat and Vulnerability Management',
    'WORKFORCE': 'Workforce Management',
}

SKIP_SHEETS = {'Home', 'E_CAT', 'G_CAT', 'L_CAT', 'Dashboard 1', 'Dashboard 2', 'Aggregate', 'Lookup'}

def parse_mil(val):
    """Convert 'MIL-1' -> 1, etc."""
    if val is None:
        return None
    m = re.match(r'MIL-(\d+)', str(val).strip())
    return int(m.group(1)) if m else None

def parse_sp(val):
    """Convert 'SP-1' -> 1, etc."""
    if val is None:
        return None
    m = re.match(r'SP-(\d+)', str(val).strip())
    return int(m.group(1)) if m else None

def get_domain_from_sheet(sheet_name):
    """Extract domain prefix from sheet name like 'ACCESS-1' -> 'ACCESS', 'THIRD-PARTIES-AP' -> 'THIRD-PARTIES'."""
    # Remove trailing -N or -AP suffix
    name = sheet_name
    if name.endswith('-AP'):
        name = name[:-3]
    else:
        # Remove trailing -<digits>
        name = re.sub(r'-\d+$', '', name)
    return name

def extract_sheet(ws, sheet_name):
    """
    Extract practices from a sheet.
    Structure (0-indexed columns):
      Col 0: practice ID or empty
      Col 3: practice text or 'Context and Guidance' label or guidance text
      Col 9: MIL value
      Col 11: SP value

    Row pattern per practice block (rows 8+ starting from row index 7):
      Row A (id_row):   col0=ID, col3=practice text, col9=MIL, col11=SP
      Row B (blank):    col0='' (blank spacer)
      Row C (label):    col0='', col3='Context and Guidance'
      Row D (guidance): col0='', col3=guidance text
      Row E (blank):    col0='' (another blank spacer before next practice)
    """
    rows = list(ws.iter_rows(values_only=True))

    # Row index 1 (0-based) = row 2 in Excel: sheet ID
    # Row index 2 (0-based) = row 3 in Excel: sub-objective name
    sub_obj_id = str(rows[1][3]).strip() if len(rows) > 1 and rows[1][3] else sheet_name
    sub_obj_name = str(rows[2][3]).strip() if len(rows) > 2 and rows[2][3] else ''

    practices = []
    i = 7  # Start from row index 7 (row 8 in Excel = first practice row)

    while i < len(rows):
        row = rows[i]
        col0 = row[0] if row[0] is not None else ''
        col3 = row[3] if len(row) > 3 and row[3] is not None else ''
        col9 = row[9] if len(row) > 9 else None
        col11 = row[11] if len(row) > 11 else None

        col0_str = str(col0).strip()
        col3_str = str(col3).strip() if col3 else ''

        # A practice ID row: col0 has the practice ID (non-empty, not just whitespace)
        if col0_str and col0_str != '' and col3_str:
            practice_id = col0_str
            practice_text = col3_str
            mil = parse_mil(col9)
            sp = parse_sp(col11)

            # Look ahead for guidance
            guidance = ''
            j = i + 1
            # Skip blank row (col0='')
            while j < len(rows):
                r = rows[j]
                r0 = str(r[0]).strip() if r[0] is not None else ''
                r3 = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ''

                # If col0 is non-empty and col3 is non-empty, it's the next practice
                if r0 and r0 != '' and r3:
                    break

                # Look for 'Context and Guidance' label row
                if r3 == 'Context and Guidance':
                    j += 1
                    # Next row is the guidance text
                    if j < len(rows):
                        r2 = rows[j]
                        r2_3 = str(r2[3]).strip() if len(r2) > 3 and r2[3] is not None else ''
                        if r2_3:
                            guidance = r2_3
                        j += 1
                    break
                j += 1

            practices.append({
                'id': practice_id,
                'text': practice_text,
                'guidance': guidance,
                'mil': mil,
                'sp': sp,
            })
            i = j  # Jump past guidance rows we already consumed
        else:
            i += 1

    return sub_obj_id, sub_obj_name, practices

def main():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # Gather domains dict: domain_id -> { name, sub_objectives: [...] }
    domains_dict = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        domain_id = get_domain_from_sheet(sheet_name)
        is_ap = sheet_name.endswith('-AP')

        ws = wb[sheet_name]
        sub_obj_id, sub_obj_name, practices = extract_sheet(ws, sheet_name)

        if domain_id not in domains_dict:
            domains_dict[domain_id] = {
                'id': domain_id,
                'name': DOMAIN_NAMES.get(domain_id, domain_id),
                'sub_objectives': [],
            }

        domains_dict[domain_id]['sub_objectives'].append({
            'id': sub_obj_id,
            'name': sub_obj_name,
            'is_anti_pattern': is_ap,
            'practices': practices,
        })

    # Convert to list, preserving insertion order (which follows sheet order)
    domains_list = list(domains_dict.values())

    output = {'domains': domains_list}

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nWrote JSON to: {OUTPUT_PATH}\n")

    # ── Summary ──────────────────────────────────────────────────────────────
    total_practices = 0
    for domain in domains_list:
        domain_count = sum(len(so['practices']) for so in domain['sub_objectives'])
        total_practices += domain_count
        print(f"Domain: {domain['id']} ({domain['name']})")
        print(f"  Sub-objectives: {len(domain['sub_objectives'])}, Total practices: {domain_count}")
        for so in domain['sub_objectives']:
            ap_flag = ' [AP]' if so['is_anti_pattern'] else ''
            print(f"    {so['id']}{ap_flag}: {len(so['practices'])} practices — {so['name']}")

    print(f"\nGRAND TOTAL practices extracted: {total_practices}")

    # ── First 5 practices per domain ─────────────────────────────────────────
    print("\n" + "="*80)
    print("FIRST 5 PRACTICES PER DOMAIN (for verification)")
    print("="*80)
    for domain in domains_list:
        print(f"\n>>> Domain: {domain['id']} — {domain['name']}")
        shown = 0
        for so in domain['sub_objectives']:
            if so['is_anti_pattern']:
                continue  # Skip AP sheets in the preview; they get shown separately
            for p in so['practices']:
                if shown >= 5:
                    break
                print(f"  [{p['id']}] MIL-{p['mil']} SP-{p['sp']}")
                print(f"    Text: {p['text'][:120]}...")
                print(f"    Guidance: {p['guidance'][:100]}..." if p['guidance'] else "    Guidance: (none)")
                shown += 1
            if shown >= 5:
                break

if __name__ == '__main__':
    main()
